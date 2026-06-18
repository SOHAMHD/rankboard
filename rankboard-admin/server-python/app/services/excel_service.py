"""EXCEL BULK IMPORT — generate the sample template, and parse uploads.

Design choices worth understanding:

- We READ uploads in read_only mode and WRITE the sample with styling.
  openpyxl is the right tool here (not pandas) because we want precise
  control over headers, validation, and a styled template — and because
  it streams large files without loading everything into memory.

- Validation is PER ROW and never trusts the file. A user's spreadsheet
  is just untrusted input wearing a friendly extension: rows can be
  blank, mistyped, duplicated, or malicious. We collect the good rows
  and a per-row reason for every bad one, then report both — so the
  user fixes 3 bad rows instead of being told only "invalid file".

- The parser returns plain dicts; the ROUTE decides what to insert.
  Keeping DB writes out of here keeps this unit testable and focused.

- The user only ever supplies the KEYWORD. A keyword's Google position
  is discovered automatically the next time "Check rankings" runs — so
  the template has a single column and the parser ignores anything else.
"""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# The contract between the sample file and the parser: this exact
# header. Keyword is the only thing we ask the user for.
COLUMNS = ["keyword"]
MAX_ROWS = 1000  # guard against someone uploading a 1M-row monster


def build_sample_workbook() -> bytes:
    """Return a styled .xlsx template as bytes, ready to stream to the
    browser. Row 1 = header, then a few example rows, then a notes
    block so the user knows the rules without reading docs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Keywords"

    header_fill = PatternFill("solid", start_color="EA580C")  # the app's orange
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    body_font = Font(name="Arial")

    # Header row
    ws.append(["Keyword"])
    cell = ws.cell(row=1, column=1)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Example rows (clearly illustrative)
    examples = [
        "online yoga classes",
        "meditation retreat rishikesh",
        "pranayama breathing course",
    ]
    for i, kw in enumerate(examples):
        ws.append([kw])
        ws.cell(row=2 + i, column=1).font = body_font

    ws.column_dimensions["A"].width = 46

    # Notes a couple of rows below the data
    notes_start = 2 + len(examples) + 1
    notes = [
        "How to use this template:",
        "• Replace the example rows above with your own keywords — one per row.",
        "• Keyword: the search term you want to track (required).",
        "• You don't need to enter any rank — RankBoard finds each keyword's current position automatically the next time you run “Check rankings”.",
        "• Keep the header row. Delete these notes if you like.",
        "• Up to %d keywords per file." % MAX_ROWS,
    ]
    for i, line in enumerate(notes):
        cell = ws.cell(row=notes_start + i, column=1)
        cell.value = line
        cell.font = Font(name="Arial", italic=True, color="78716C", bold=(i == 0))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_keyword_workbook(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    """Parse an uploaded .xlsx. Returns (valid_rows, errors).

    valid_rows: [{"term"}, ...]   — rank columns are intentionally ignored
    errors:     [{"row": <excel row number>, "reason": "..."}, ...]
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError("That file couldn't be read as an Excel (.xlsx) workbook.")

    ws = wb.active
    valid: list[dict] = []
    errors: list[dict] = []
    seen_terms: set[str] = set()
    data_rows = 0

    # enumerate starting at row 1; we skip row 1 as the header
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row == 1:
            continue  # header
        if row is None:
            continue

        term_raw = row[0] if len(row) > 0 else None

        # Blank keyword → silently skip (trailing empties, notes, etc.).
        # Keyword is the only field, so a row with no keyword is just noise.
        if term_raw is None or str(term_raw).strip() == "":
            continue

        # Skip the template's own notes block if left in.
        if isinstance(term_raw, str) and term_raw.strip().startswith(("How to use", "•")):
            continue

        data_rows += 1
        if data_rows > MAX_ROWS:
            errors.append({"row": excel_row, "reason": f"File exceeds the {MAX_ROWS}-keyword limit; remaining rows ignored."})
            break

        term = str(term_raw).strip().lower()

        # Duplicate within the same file → keep the first, flag the rest.
        if term in seen_terms:
            errors.append({"row": excel_row, "reason": f"Duplicate of an earlier row for “{term}”; skipped."})
            continue
        seen_terms.add(term)

        valid.append({"term": term})

    return valid, errors
