import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

COLUMNS = ["keyword"]
MAX_ROWS = 1000

#: Matches routers/projects.py MAX_TERM_LEN. A cell holding a paragraph is a
#: mis-shaped spreadsheet, not a long-tail keyword.
MAX_TERM_LEN = 200


def build_sample_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Keywords"

    header_fill = PatternFill("solid", start_color="EA580C")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    body_font = Font(name="Arial")

    ws.append(["Keyword"])
    cell = ws.cell(row=1, column=1)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    examples = [
        "online yoga classes",
        "meditation retreat rishikesh",
        "pranayama breathing course",
    ]
    for i, kw in enumerate(examples):
        ws.append([kw])
        ws.cell(row=2 + i, column=1).font = body_font

    ws.column_dimensions["A"].width = 46

    notes_start = 2 + len(examples) + 1
    # These notes used to promise that the app "finds each keyword's current
    # position automatically the next time you run Check rankings". There is no
    # such feature and never was — every rank in this app is typed in by hand —
    # so the template was describing a workflow the product doesn't have.
    notes = [
        "How to use this template:",
        "• Replace the example rows above with your own keywords — one per row.",
        "• Keyword: the search term you want to track (required).",
        "• This file adds the keywords only. Enter each month's position afterwards in the Keywords grid — ranks are recorded by hand.",
        "• Duplicates are skipped, so it's safe to re-import a file you've added to.",
        "• Keep the header row. Delete these notes if you like.",
        "• Up to %d keywords per file." % MAX_ROWS,
    ]
    for i, line in enumerate(notes):
        cell = ws.cell(row=notes_start + i, column=1)
        cell.value = line
        cell.font = Font(name="Arial", italic=True, color="78716C", bold=(i == 0))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


#: Read at most this many rows off the sheet. A workbook whose stored dimension
#: is wrong — or which has a stray format applied a million rows down — reports a
#: max_row far past the last real value, and read_only iteration walks every one
#: of them. MAX_ROWS is the keyword limit; this is the hard stop on how far we
#: look for them, generous enough to survive a file with gaps and blank rows.
_MAX_SCAN_ROWS = 50_000


def _iter_terms(ws) -> list[tuple[int, object]]:
    """Materialise (row number, first cell) pairs.

    Returned as a list rather than a generator so that a sheet which fails
    mid-parse fails inside the caller's try block. read_only iteration is lazy:
    as a generator the exception surfaced later, in the middle of the validation
    loop, where nothing was catching it.
    """
    out: list[tuple[int, object]] = []
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row > _MAX_SCAN_ROWS:
            break
        out.append((excel_row, row[0] if row else None))
    return out


def parse_keyword_workbook(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError("That file couldn't be read as an Excel (.xlsx) workbook.")

    valid: list[dict] = []
    errors: list[dict] = []
    seen_terms: set[str] = set()
    data_rows = 0

    try:
        # `wb.active` is None when the workbook records no active sheet — rare
        # from Excel itself, common from exporters and "save as xlsx" tools.
        # Falling back to the first sheet beats an AttributeError deep in the
        # loop, which surfaced as a bare 500 with no hint that the file was the
        # problem.
        ws = wb.active
        if ws is None:
            ws = next(iter(wb.worksheets), None)
        if ws is None:
            raise ValueError("That workbook has no sheets to read.")

        rows = _iter_terms(ws)
    except ValueError:
        wb.close()
        raise
    except Exception:
        # read_only parsing is lazy: a malformed sheet doesn't fail in
        # load_workbook above, it fails here, mid-iteration. Everything past this
        # point used to escape as an unhandled 500.
        wb.close()
        raise ValueError("That file couldn't be read as an Excel (.xlsx) workbook.")

    # read_only mode holds the workbook's zip open. Every import leaked one file
    # handle until the process recycled.
    wb.close()

    for excel_row, term_raw in rows:
        if excel_row == 1:
            continue

        if term_raw is None or str(term_raw).strip() == "":
            continue

        if isinstance(term_raw, str) and term_raw.strip().startswith(("How to use", "•")):
            continue

        data_rows += 1
        if data_rows > MAX_ROWS:
            errors.append({"row": excel_row, "reason": f"File exceeds the {MAX_ROWS}-keyword limit; remaining rows ignored."})
            break

        term = str(term_raw).strip().lower()

        if len(term) > MAX_TERM_LEN:
            errors.append({
                "row": excel_row,
                "reason": f"Longer than {MAX_TERM_LEN} characters — is this a keyword? Skipped.",
            })
            continue

        if term in seen_terms:
            errors.append({"row": excel_row, "reason": f"Duplicate of an earlier row for “{term}”; skipped."})
            continue
        seen_terms.add(term)

        valid.append({"term": term})

    return valid, errors
